
from student import Student
from colleague import Colleague
from friend import Friend
from relative import Relative
import openpyxl
from openpyxl.styles import Font
import smtplib
# from playsound import playsound


def add_contact(contact_type, name, birthday, phone, email,height, weight, additional_info=None):
    """
    Adds a new contact to the address book.

    contact_type: Type of the contact (student, colleague, friend, relative)
    name: Name of the person
    birthday: Birthday of the person
    phone: Phone number of the person
    email: Email address of the person
    additional_info: Additional information relevant to the contact type
    """
    # Check if the contact already exists in the Excel file
    filename = "contacts.xlsx"
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Contacts']
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[:7] == (name, birthday, phone, email, height, weight,contact_type):
                # Contact already exists, return the existing contact details
                return {'Status': 'Existing', 'Details': row}

        # If the contact does not exist, add a new contact
        if contact_type == 'student':
            contact = Student(name, birthday, phone, email,height, weight, *additional_info) #*符号用于函数调用时解包列表或元组,高级！
        elif contact_type == 'colleague':
            contact = Colleague(name, birthday, phone, email,height, weight, *additional_info)
        elif contact_type == 'friend':
            contact = Friend(name, birthday, phone, email,height, weight, additional_info)
        elif contact_type == 'relative':
            contact = Relative(name, birthday, phone, email, height, weight,additional_info)
        else:
            raise ValueError("Invalid contact type. Must be one of: student, colleague, friend, relative.")

        contact.write_to_excel()
        return {'Status': 'Added', 'Details': (name, birthday, phone, email, height, weight,contact_type)}
    except FileNotFoundError:
        # File does not exist, proceed to add a new contact
        pass

    # Add a new contact if the file does not exist or the contact is new
    if contact_type == 'student':
        contact = Student(name, birthday, phone, email, height, weight,*additional_info)
    elif contact_type == 'colleague':
        contact = Colleague(name, birthday, phone, email,height, weight, *additional_info)
    elif contact_type == 'friend':
        contact = Friend(name, birthday, phone, email,height, weight, additional_info)
    elif contact_type == 'relative':
        contact = Relative(name, birthday, phone, email,height, weight, additional_info)
    else:
        raise ValueError("Invalid contact type. Must be one of: student, colleague, friend, relative.")

    contact.write_to_excel()
    # playsound("voice/add_contact.mp3")
    print('Added')
    #sync
    sync()
    return {'Status': 'Added', 'Details': (name, birthday, phone, email,height, weight, contact_type)}

# print('student', "Alice", "1995-02-10", "123-456-7890", "alice@example.com", 165, 55, ["MIT", "Senior", "Computer Science", 3.8])
# print('student', "Bob", "1996-08-15", "987-654-3210", "bob@example.com", 175, 70, ["Stanford", "Junior", "Physics", 3.6]))
# Creating 10 objects with the specified parameter order in the constructors
# Using fictional data for demonstration

# Creating 2 Student objects
# add_contact('student', "Alice", "1995-02-10", "123-456-7890", "alice@example.com", 165, 55, ["MIT", "Senior", "Computer Science", 3.8])
# add_contact('student', "Bob", "1996-08-15", "987-654-3210", "bob@example.com", 175, 70, ["Stanford", "Junior", "Physics", 3.6])
#
# # Creating 2 Colleague objects
# add_contact('colleague', "Charlie", "1985-03-20", "555-666-7777", "charlie@example.com", 180, 75, ["TechCorp", 80000])
# add_contact('colleague', "Diana", "1990-07-22", "444-333-2222", "diana@example.com", 170, 65, ["FinanceInc", 95000])
#
# # Creating 3 Friend objects
# add_contact('friend', "Ethan", "1992-12-01", "222-333-4444", "ethan@example.com", 178, 72, ["Met at a conference"])
# add_contact('friend', "Fiona", "1993-11-03", "111-222-3333", "fiona@example.com", 165, 55, ["Childhood friend"])
# add_contact('friend', "George", "1994-10-05", "333-444-5555", "george@example.com", 182, 80, ["University roommate"])
#
# # Creating 3 Relative objects
# add_contact('relative', "Hannah", "1975-04-09", "999-888-7777", "hannah@example.com", 160, 60, ["Cousin"])
# add_contact('relative', "Ian", "1980-05-12", "888-777-6666", "ian@example.com", 185, 85, ["Uncle"])
# add_contact('relative', "Julia", "1983-06-14", "777-666-5555", "julia@example.com", 168, 58, ["Sister"])








#creat object lazily

import random

def create_mixed_contacts(num_objects):
    """
    Creates a specified number of mixed type contacts (Student, Colleague, Friend, Relative)
    with partially similar information.

    :param num_objects: Number of objects to create
    :return: List of created objects of mixed types
    """
    contacts = []
    names = ["Alice", "Bob", "Charlie", "Diana", "Ethan", "Fiona", "George", "Hannah", "Ian", "Julia"]
    types = ["student", "colleague", "friend", "relative"]
    relationships = ["Sister", "Brother", "Cousin", "Uncle", "Aunt"]
    colleges = ["MIT", "Stanford", "Harvard"]
    companies = ["TechCorp", "FinanceInc", "HealthCo"]

    for _ in range(num_objects):
        contact_type = random.choice(types)
        name = random.choice(names)
        birthday = f"198{random.randint(0, 9)}-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}"
        phone = f"{random.randint(100, 999)}-{random.randint(100, 999)}-{random.randint(1000, 9999)}"
        email = f"{name.lower()}@example.com"
        height = random.randint(150, 200)
        weight = random.randint(50, 100)

        if contact_type == "student":
            college = random.choice(colleges)
            grade = random.choice(["Freshman", "Sophomore", "Junior", "Senior"])
            major = random.choice(["Computer Science", "Physics", "Engineering"])
            gpa = round(random.uniform(2.0, 4.0), 2)
            add_contact( "student",name, birthday, phone, email, height, weight, [college, grade, major, gpa])
        
        elif contact_type == "colleague":
            company = random.choice(companies)
            income = random.randint(50000, 120000)
            add_contact("colleague", name, birthday, phone, email, height, weight, [company, income])

        elif contact_type == "friend":
            acquaintance_info = random.choice(["Met at a conference", "Childhood friend", "University roommate"])
            add_contact("friend",name, birthday, phone, email, height, weight, [acquaintance_info])

        elif contact_type == "relative":
            relationship = random.choice(relationships)
            add_contact( "relative",name, birthday, phone, email, height, weight, [relationship])

        



# Creating 100 mixed contact objects
# mixed_contacts = create_mixed_contacts(100)


def delete_contact(name):
    filename = "contacts.xlsx"
    try:
        # Load the workbook and get the active sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Contacts']

        # Iterate through the rows to find the contact
        row_to_delete = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == name:  # Check if the name matches
                row_to_delete = row[0].row  # Get the row number
                break

        if row_to_delete:
            # Delete the contact row
            sheet.delete_rows(row_to_delete)
            wb.save(filename)
            #sync
            sync()
            # playsound("voice/delete_contact.mp3")
            return "Contact deleted successfully."
        else:
            # Contact not found
            return "No such person found."
    except FileNotFoundError:
        # If the file does not exist
        return "No such person found."

# Example usage:
# result = delete_contact("John Doe")
# print(result)


def modify_contact(name, info_type, new_value):
    filename = "contacts.xlsx"
    try:
        # Load the workbook and get the active sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Contacts']

        # Define a dictionary to map info types to column indices
        info_type_to_column = {
            'birthday': 2,
             'phone': 3,
             'email': 4,
             'height': 5, 
             'weight': 6,
             'Contact Type': 7,
             'company': 5, 
             'relationship': 8 , 
             'grade': 8 ,#student
             'gpa': 8 ,#student
             'company': 8 ,#colleague
             'income': 8 ,#colleague
             'acquaintance_info': 8 ,#friend
             'relationship': 8 ,#relative
             'major': 8 #student
        }

        # Check if the info type is valid
        if info_type not in info_type_to_column:
            return "Invalid information type."

        # Iterate through the rows to find the contact
        contact_found = False
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == name:  # Check if the name matches
                contact_found = True
                # Modify the specified information
                column_index = info_type_to_column[info_type]
                sheet.cell(row=row[0].row, column=column_index, value=new_value)
                break

        if contact_found:
            wb.save(filename)
            # playsound("voice/modify_contact.mp3")
            #sync
            sync()
            return f"Contact '{name}' information '{info_type}' updated to '{new_value}'."
        else:
            return "No such person found."
    except FileNotFoundError:
        # If the file does not exist
        return "No such person found."


# print(modify_contact('John Doe','birthday','2077-01-01'))
# print(modify_contact('a','birthday','2077-01-01'))


def search_contact(search_info):
    filename = "contacts.xlsx"
    try:
        # Load the workbook and get the active sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Contacts']
        found_contacts = []

        # Iterate through the rows to search for the contact
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if any(search_info in str(cell) for cell in row):
                # Add the contact information to the list if any cell partially matches
                found_contacts.append(row)

        num = len(found_contacts)
        if found_contacts:
            # Return the number of contacts found along with their details
            return {'Status': f'{num} Contacts Found', 'Details': found_contacts}
        else:
            return "No such person found."
    except FileNotFoundError:
        # If the file does not exist
        return "No such person found."
    
# print(search_contact("John Doe"))
# print(search_contact("1976-05-05"))
# print(search_contact("henry.wilson@email.com"))
# print(search_contact("colleague"))
# print(search_contact("John"))



def sort(sort_key):
    """
    Sorts all contacts from a contacts.xlsx file based on the specified key.
    
    """
    if sort_key not in ['birthday', 'height', 'weight', 'phone']:
        raise ValueError("Invalid sort key. Must be one of: 'birthday', 'height', 'weight', 'phone'.")

    filename = "contacts.xlsx"
    try:
        # Load the workbook and get the active sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Contacts']
        contacts = []

        # Iterate through the rows to collect all contacts
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            # Extracting contact information
            name, birthday, phone, email, height, weight = row[:6]
            contact = {
                'name': name, 
                'birthday': birthday,
                'phone': phone,
                'email': email,
                'height': height,
                'weight': weight
            }
            contacts.append(contact)

        # Sorting logic
        if sort_key in ['height', 'weight']:
            sorted_contacts = sorted(contacts, key=lambda contact: float(contact[sort_key]))
        else:  # For 'birthday' and 'phone', sorting as strings
            sorted_contacts = sorted(contacts, key=lambda contact: contact[sort_key])
        # playsound("voice/sort.mp3")
        return sorted_contacts

    except FileNotFoundError:
        return "contacts.xlsx file not found."
# print(sort("height"))


def sort_students(sort_key):

    if sort_key not in ['grade', 'gpa']:
        raise ValueError("Invalid sort key. Must be one of: 'grade', 'gpa'.")

    filename = "contacts.xlsx"
    try:
        # Load the workbook and get the active sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Contacts']
        students = []

        # Iterate through the rows to collect Student objects
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            contact_type = row[6]  # Assuming contact type is in the 6th column
            if contact_type == 'student':
                # Extracting Student information
                name, birthday, phone, email, height, weight, additional_info = row[:5] + row[6:]
                college, grade, major, gpa = additional_info[1:-1].split(', ')
                student = Student(name, birthday, phone, email, height, weight, *[college, grade, major, gpa])
                students.append(student)

        # Sorting logic
        if sort_key == 'grade':
            sorted_students = sorted(students, key=lambda student: student.grade)

            result=[]
            for x in sorted_students:
                result.append((x.name,f"grade{x.grade}"))
            # playsound("voice/sortall.mp3")
            return result
   
        elif sort_key == 'gpa':
            sorted_students = sorted(students, key=lambda student: student.gpa, reverse=True)  

            result=[]
            for x in sorted_students:
                result.append((x.name,f"gpa:{x.gpa}"))
            # playsound("voice/sortall.mp3")
            return result

    except FileNotFoundError:
        return "contacts.xlsx file not found."

# Example usage:
#sorted_students_by_grade = sort_students('grade')
# sorted_students_by_gpa = sort_students('gpa')
# print(sorted_students_by_grade)
# print(sorted_students_by_gpa)



def sort_colleagues_by_income():


    filename = "contacts.xlsx"
    try:
        # Load the workbook and get the active sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Contacts']
        colleagues = []

        # Iterate through the rows to collect Student objects
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            contact_type = row[6]  # Assuming contact type is in the 6th column
            if contact_type == 'colleague':
                # Extracting colleague information
                name, birthday, phone, email, height, weight, additional_info = row[:5] + row[6:]
                company, income = additional_info[1:-1].split(', ')
                income = float(income)  # Convert income to a float
                colleague = Colleague(name, birthday, phone, email, height, weight, *[company, income])
                colleagues.append(colleague)

        # Sorting colleagues by income
        sorted_colleagues = sorted(colleagues, key=lambda colleague: colleague.income, reverse=True)

        result=[]
        for x in sorted_colleagues:
            result.append((x.name,"income",x.income))
        # playsound("voice/sortall.mp3")
        return result

    except FileNotFoundError:
        return "contacts.xlsx file not found."
    
# sorted_colleagues = sort_colleagues_by_income()
# print(sorted_colleagues)





# Writing a function to filter contacts based on various criteria from the contacts.xlsx file.


def filter_contacts(criteria):
    filename = "contacts.xlsx"
    try:
        # Load the workbook and get the active sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active

        # Check if there are contacts in the sheet
        if sheet.max_row < 2:
            return "No contacts available."

        filtered_contacts = []

        # Define column headers to column index mapping
        column_headers = {'Name': 0, 'Birthday': 1, 'Phone': 2, 'Email': 3, 'Height': 4, 'Weight': 5, 'Contact Type': 6,
                          'Additional Info': 7}

        # Iterate through the rows to filter contacts
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            match = True
            for key, value in criteria.items():

                if key in column_headers:

                    cell_value = row[column_headers[key]]

                    if type(cell_value) == None:
                        return "no contacts available"
                    if isinstance(value, tuple):
                        operator, compare_value = value
                        if operator == '>' and not (cell_value > compare_value):
                            match = False
                            break
                        elif operator == '<' and not (cell_value < compare_value):
                            match = False
                            break
                    else:
                        if str(cell_value) != str(value):
                            match = False
                            break
            if match:
                filtered_contacts.append(row)

        return filtered_contacts

    except FileNotFoundError:
        return "contacts.xlsx file not found."

# Example usage:
# criteria = {"Height": ('>', 170),    "Weight": ('<', 70)}
# filtered_contacts = filter_contacts(criteria)


# This function assumes that the column names in the Excel file match the keys in the 'criteria' dictionary.
# Adjust the column_headers mapping if your Excel file format is different.


def filter_contacts_advanced(criteria):
    """
    Filters contacts based on the given criteria from the contacts.xlsx file, including greater than or less than conditions.

    :param criteria: A dictionary where keys are column names and values are the criteria for filtering.
                     Use tuples for greater than or less than conditions, e.g., ('>', 170) for height greater than 170.
    :return: List of contacts that meet the criteria
    """
    filename = "contacts.xlsx"
    try:
        # Load the workbook and get the active sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Contacts']
        filtered_contacts = []

        # Define column headers to column index mapping
        column_headers = {header.value: idx + 1 for idx, header in enumerate(sheet[1])}

        # Iterate through the rows to filter contacts
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            match = True
            for key, value in criteria.items():
                if key in column_headers:
                    cell_value = row[column_headers[key] - 1]
                    if isinstance(value, tuple):
                        operator, compare_value = value
                        if operator == '>' and not (cell_value > compare_value):
                            match = False
                            break
                        elif operator == '<' and not (cell_value < compare_value):
                            match = False
                            break
                    else:
                        if str(cell_value) != str(value):
                            match = False
                            break
            if match:
                filtered_contacts.append(row)
        # playsound("voice/filter_contacts.mp3")
        return filtered_contacts

    except FileNotFoundError:
        return "contacts.xlsx file not found."
# criteria = {'Height': ('>', 170), 'Weight': ('<', 70)}
# filtered_contacts = filter_contacts_advanced(criteria)
# print(filtered_contacts)
# This function now handles conditions for greater than or less than for fields like height and weight.

import matplotlib.pyplot as plt

def fig_Contact():
    filename = "contacts.xlsx"
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Contacts']

        dict_contact={}
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            contact_type = row[6]
            if contact_type not in dict_contact.keys():
                dict_contact[contact_type]=1
            else:
                dict_contact[contact_type]+=1
        lst_temp=[]
        for key in dict_contact.keys():
            num=dict_contact[key]
            lst_temp.append([key,num])
        lst_temp.sort(key=lambda ele: ele[1], reverse=True)
        lst_contact=[]
        lst_number=[]
        for element in lst_temp:
            lst_contact.append(element[0])
            lst_number.append(element[1])
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.pie(lst_number,
                labels=lst_contact,  # 设置饼图标签
                autopct="(%1.1f%%)",  # 饼块内标签。
                startangle=90
                )
        plt.title("通讯录关系占比")
        plt.show()

    except FileNotFoundError:
        return "contacts.xlsx file not found."
# Example usage:
# fig_Contact()




def depulicate_check(): # not nessasary
    import pandas as pd

    # Define the paths for the input and output files
    input_file_path = 'contacts.xlsx'
    output_file_path = 'contacts.xlsx'

    # Read the Excel file
    # Assuming that your data is in the first sheet of the Excel file
    df = pd.read_excel(input_file_path)

    # Print the number of rows in the original data for reference
    print(f"Number of rows in original data: {len(df)}")

    # Remove duplicate rows
    # The drop_duplicates method keeps the first occurrence of each set of duplicate rows
    df_deduplicated = df.drop_duplicates()

    # Print the number of rows after removing duplicates
    print(f"Number of rows after removing duplicates: {len(df_deduplicated)}")

    # Save the deduplicated data to a new Excel file
    # index=False means do not include the row index when saving
    df_deduplicated.to_excel(output_file_path, index=False)

    print(f"Deduplicated data saved to {output_file_path}")
# depulicate_check()

def sync():
    from bypy import ByPy
    bp = ByPy()

    bp.upload(
    r"contacts.xlsx"
    )
    print(bp.list())
    return
# sync()

# Writing a function to return all contacts from the contacts.xlsx file.

def get_all_contacts():
    """
    Returns all contacts from the contacts.xlsx file.

    :return: List of all contacts
    """
    filename = "contacts.xlsx"
    try:
        # Load the workbook and get the active sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb['Contacts']
        all_contacts = []

        # Iterate through the rows to collect all contacts
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            all_contacts.append(row)

        return all_contacts

    except FileNotFoundError:
        return "contacts.xlsx file not found."

# Example usage:
# contacts = get_all_contacts()

# This function simply reads and returns all rows from the contacts.xlsx file.\
    

def clear():
    filename = "contacts.xlsx"
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook['Contacts']

        # 获取标题行
        header = sheet[1]

        # 删除除标题行外的所有行
        sheet.delete_rows(2, sheet.max_row)

        # 保存并关闭文件
        workbook.save(filename)
        workbook.close()
        print("All contacts have been cleared.")

        #添加实例联系人保证通讯录中有至少一个联系人
        add_contact('friend', "yourself", "2000-11-03", "111-222-3333", "yourself@example.com", 165, 55, ["Yourself"])
    except Exception as e:
        print(f"Error occurred while clearing contacts: {str(e)}")