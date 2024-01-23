
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from openpyxl.styles import Font
import os
import datetime


class Person:
    def __init__(self, name, birthday, phone, email, height, weight, contact_type, additional_info=None):
        # 初始化 Person 类的实例
        self.name = name  # 姓名
        self.birthday = birthday  # 出生日期
        self.phone = phone  # 电话号码
        self.email = email  # 电子邮件地址
        self.contact_type = contact_type  # 联系方式类型
        self.height = height  # 身高（单位：厘米）
        self.weight = weight  # 体重（单位：千克）
        self.additional_info = additional_info  # 额外信息，可选，默认为 None

    def __str__(self):
        # 返回 Person 类的字符串表示
        return f"Person(name='{self.name}', email='{self.email}')"

    def __repr__(self):
        # 返回 Person 类的正式表示
        return f"Person(name='{self.name}', email='{self.email}')"

    def __eq__(self, other):
        # 比较两个 Person 实例是否相等
        if isinstance(other, Person):
            return self.name == other.name and self.email == other.email
        return False

    def __hash__(self):
        # 计算 Person 实例的哈希值
        return hash((self.name, self.email))

    def calculate_bmi(self):
        # 计算BMI（Body Mass Index）
        bmi = (self.weight / (self.height / 100) ** 2)
        return round(bmi, 2)

    def get_age(self):
        # 计算年龄
        birth_year = int(self.birthday.split("-")[0])
        current_year = datetime.datetime.now().year
        age = current_year - birth_year
        return age

    def send_email(self, message):
        # 发送电子邮件
        print(f"Sending email to {self.email}: {message}")

    def show_friends(self):
        # 查看朋友列表
        return [friend.name for friend in self.friends]

    def add_additional_info(self, key, value):
        # 添加额外信息
        if self.additional_info is None:
            self.additional_info = {}
        self.additional_info[key] = value

    def get_additional_info(self, key):
        # 获取额外信息
        if self.additional_info and key in self.additional_info:
            return self.additional_info[key]
        else:
            return None



    def write_to_excel(self, filename="contacts.xlsx"):
        # Check if the Excel file already exists
        if os.path.exists(filename):
            # Open the existing workbook
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active
        else:
            # Create a new workbook
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Contacts'

            # Adding headers if the file is new
            headers = ["Name", "Birthday", "Phone", "Email","height", "weight", "Contact Type", "Additional Info"]
            if sheet.max_row == 1:
                for col, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
    
        # Add contact information along with contact type and specific attributes
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=self.name)
        sheet.cell(row=row, column=2, value=self.birthday)
        sheet.cell(row=row, column=3, value=self.phone)
        sheet.cell(row=row, column=4, value=self.email)
        sheet.cell(row=row, column=5, value=self.height)
        sheet.cell(row=row, column=6, value=self.weight)
        sheet.cell(row=row, column=7, value=self.contact_type)
        sheet.cell(row=row, column=8, value=str(self.additional_info))

        # Save the workbook
        wb.save(filename)


    # def add_contact(self, contact_name):
    #     if contact_name not in self.contacts:
    #         self.contacts.append(contact_name)
    #         return
    #     else:
    #         print("已有联系人")
    #         return

    # def delete_contact(self, contact_name):
    #     for contact in self.contacts:
    #         if contact.name == contact_name:
    #             self.contacts.remove(contact)
    #         else:
    #             print("无此联系人")
    #         return

    # def send_birthday_email(self):
    #     sender_email = 'your_email@example.com'  # Replace with your email
    #     sender_password = 'your_password'        # Replace with your password
    #     smtp_server = 'smtp.example.com'         # Replace with your SMTP server
    #     smtp_port = 587
    #
    #     subject = '生日快乐！'
    #     body = f'亲爱的 {self.name}，祝你生日快乐！'
    #
    #     message = MIMEMultipart()
    #     message['From'] = sender_email
    #     message['To'] = self.email
    #     message['Subject'] = subject
    #     message.attach(MIMEText(body, 'plain'))
    #
    #     try:
    #         server = smtplib.SMTP(smtp_server, smtp_port)
    #         server.starttls()
    #         server.login(sender_email, sender_password)
    #         server.sendmail(sender_email, self.email, message.as_string())
    #         server.quit()
    #         print(f'生日祝福邮件已发送至 {self.email}')
    #     except Exception as e:
    #         print(f'邮件发送失败：{str(e)}')

    # Other methods and logic (same as before)


# 示例用法
# person1 = Person("John", "1990-01-01", "123-456-7890", "john@example.com", 180, 75, "friend")
# person2 = Person("Jane", "1992-03-15", "987-654-3210", "jane@example.com", 165, 55, "friend")
# person3 = Person("Alice", "1988-07-20", "555-555-5555", "alice@example.com", 170, 60, "colleague")

# person1.add_friend(person2)
# person1.add_friend(person3)

# print(person1.show_friends())  # 输出: ['Jane', 'Alice']

# person1.add_additional_info("Hobbies", ["Reading", "Swimming"])
# person1.add_additional_info("Education", "Bachelor's Degree")

# print(person1.get_additional_info("Hobbies"))  # 输出: ['Reading', 'Swimming']
# print(person1.get_additional_info("Education"))  # 输出: Bachelor's Degree
        