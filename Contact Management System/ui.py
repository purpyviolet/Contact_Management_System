import sys

from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QTextEdit, QMessageBox, QRadioButton,QGroupBox,QFormLayout,QComboBox,QSplitter, QSizePolicy
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView
from address_book import add_contact, search_contact, sort, delete_contact, modify_contact,sort_students,sort_colleagues_by_income,filter_contacts,filter_contacts_advanced,get_all_contacts, clear  # Import your functions
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5.QtWidgets import QVBoxLayout
import openpyxl
from find_send_email import get_tomorrows_birthdays, send_birthday_email, birthday_wishes
from playsound import playsound
import random
# 首先需要一个主页面，主页面包含了系统名称
# 考虑到系统要求，这里会写这些交互在主页面
# 1显示所有信息
# 增加
# 修改
# 删除
# 查找（分多种方式）
# 检查有没有人过生日

# 排序
# 排序分为几个方面
# 首先是普通sort（根据身高体重年龄等方式sort）
# 其次是学生的sort（根据gpa sort）
# 还有就是员工的sort（根据收入sort）

# 筛选
# 筛选可以通过名字、出生年月、身高体重（比较值）来做到
# 自动上传
# 画图功能

class ContactManagementApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Contact Management System")
        self.setGeometry(100, 100, 1200, 900)

        self.tabWidget = QTabWidget(self)
        self.setCentralWidget(self.tabWidget)

        # Initialize tabs
        self.initDisplayAllContactsTab()
        self.initAddContactTab()
        self.initSearchContactsTab()
        self.initSendBirthdayEmailsTab()
        self.initModifyContactsTab()
        self.initDeleteContactsTab()
        self.initSortContactsTab()
        self.initFilterContactsTab()
        # self.initFilterContactsAdvancedTab()
        self.initFigueContactsTab()
    def initAddContactTab(self):
        self.addContactTab = QWidget()
        self.tabWidget.addTab(self.addContactTab, "Add Contacts")

        layout = QVBoxLayout()
        self.nameInput = QLineEdit()
        self.birthdayInput = QLineEdit()
        self.phoneInput = QLineEdit()
        self.emailInput = QLineEdit()
        self.heightInput = QLineEdit()
        self.weightInput = QLineEdit()
        layout.addWidget(QLabel("Name"))
        layout.addWidget(self.nameInput)
        layout.addWidget(QLabel("Birthday"))
        layout.addWidget(self.birthdayInput)
        layout.addWidget(QLabel("Phone"))
        layout.addWidget(self.phoneInput)
        layout.addWidget(QLabel("Email"))
        layout.addWidget(self.emailInput)
        layout.addWidget(QLabel("Height"))
        layout.addWidget(self.heightInput)
        layout.addWidget(QLabel("Weight"))
        layout.addWidget(self.weightInput)
        # 创建 Radio Buttons 用于选择联系人类型
        self.studentRadio = QRadioButton("Student")
        self.colleagueRadio = QRadioButton("Colleague")
        self.friendRadio = QRadioButton("Friend")
        self.relativeRadio = QRadioButton("Relative")

        # 将 Radio Buttons 放入 GroupBox 中
        self.contactTypeGroup = QGroupBox("Contact Type")
        contactTypeLayout = QHBoxLayout()
        contactTypeLayout.addWidget(self.studentRadio)
        contactTypeLayout.addWidget(self.colleagueRadio)
        contactTypeLayout.addWidget(self.friendRadio)
        contactTypeLayout.addWidget(self.relativeRadio)
        self.contactTypeGroup.setLayout(contactTypeLayout)

        # 额外的输入字段，仅在选择了特定的联系人类型时显示
        self.additionalInfoLayout = QFormLayout()
        self.collegeInput = QLineEdit()
        self.gradeInput = QLineEdit()
        self.majorInput = QLineEdit()
        self.gpaInput = QLineEdit()
        self.companyInput = QLineEdit()
        self.incomeInput = QLineEdit()
        self.acquaintanceInfoInput = QLineEdit()
        self.relationshipInput = QLineEdit()

        # 连接 Radio Buttons 的信号到槽函数
        self.studentRadio.toggled.connect(self.onContactTypeChanged)
        self.colleagueRadio.toggled.connect(self.onContactTypeChanged)
        self.friendRadio.toggled.connect(self.onContactTypeChanged)
        self.relativeRadio.toggled.connect(self.onContactTypeChanged)

        layout.addWidget(QLabel("Name"))
        layout.addWidget(self.nameInput)
        # ... 其他通用字段
        layout.addWidget(self.contactTypeGroup)
        layout.addLayout(self.additionalInfoLayout)  # 添加额外信息的布局

        addBtn = QPushButton("Add Contact")
        addBtn.clicked.connect(self.addContact)
        layout.addWidget(addBtn)

        self.addContactTab.setLayout(layout)

    def onContactTypeChanged(self):
        # 清除额外信息布局中的所有字段
        while self.additionalInfoLayout.count():
            child = self.additionalInfoLayout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        # 根据选中的联系人类型添加相应的输入字段
        if self.studentRadio.isChecked():
            self.additionalInfoLayout.addRow("College", self.collegeInput)
            self.additionalInfoLayout.addRow("Grade", self.gradeInput)
            self.additionalInfoLayout.addRow("Major", self.majorInput)
            self.additionalInfoLayout.addRow("GPA", self.gpaInput)
        elif self.colleagueRadio.isChecked():
            self.additionalInfoLayout.addRow("Company", self.companyInput)
            self.additionalInfoLayout.addRow("Income", self.incomeInput)
        elif self.friendRadio.isChecked():
            self.additionalInfoLayout.addRow("Acquaintance Info", self.acquaintanceInfoInput)
        elif self.relativeRadio.isChecked():
            self.additionalInfoLayout.addRow("Relationship", self.relationshipInput)

    def addContact(self):
        """
        添加新的联系人信息到系统中。
        该方法首先从输入字段获取联系人的基本信息，如姓名、生日、电话等。
        然后，根据用户选择的联系人类型（如学生、同事、朋友、亲戚）获取更具体的信息。
        最后，调用 add_contact 函数添加联系人，并处理响应结果。
        """

        # 从输入字段中获取联系人的基本信息
        name = self.nameInput.text()
        birthday = self.birthdayInput.text()
        phone = self.phoneInput.text()
        email = self.emailInput.text()

        # 如果用户没有输入身高或体重，将这些字段设置为 "None"
        height = self.heightInput.text() or "None"
        weight = self.weightInput.text() or "None"

        # 根据用户选择的联系人类型获取额外信息
        if self.studentRadio.isChecked():
            contact_type = "student"
            college = self.collegeInput.text()
            grade = self.gradeInput.text()
            major = self.majorInput.text()
            gpa = self.gpaInput.text()
            additional_info = [college, grade, major, gpa]
        elif self.colleagueRadio.isChecked():
            contact_type = "colleague"
            company = self.companyInput.text()
            income = self.incomeInput.text()
            additional_info = [company, income]
        elif self.friendRadio.isChecked():
            contact_type = "friend"
            acquaintance_info = self.acquaintanceInfoInput.text()
            additional_info = [acquaintance_info]
        elif self.relativeRadio.isChecked():
            contact_type = "relative"
            relationship = self.relationshipInput.text()
            additional_info = [relationship]
        else:
            # 如果没有选择联系人类型，则显示警告信息
            QMessageBox.warning(self, "Contact Type Error", "Please select a contact type.")
            return

        # 尝试将身高和体重的字符串转换为浮点数
        try:
            height = float(height) if height != "None" else None
            weight = float(weight) if weight != "None" else None
        except ValueError:
            # 如果输入的身高或体重不是有效数字，则显示错误信息
            QMessageBox.critical(self, "Input Error", "Please enter valid numbers for height and weight.")
            return

        # 调用后端函数添加新的联系人
        try:
            result = add_contact(contact_type, name, birthday, phone, email, height, weight, additional_info)
            # 根据后端函数返回的结果显示相应的消息
            if result['Status'] == 'Added':
                # 构建显示的详细信息
                detailed_info = f"Contact added successfully:\nName: {name}\nBirthday: {birthday}\nPhone: {phone}\nEmail: {email}\nHeight: {height or 'N/A'}\nWeight: {weight or 'N/A'}\nType: {contact_type.capitalize()}\nAdditional Info: {', '.join(additional_info) if additional_info else 'N/A'}"
                QMessageBox.information(self, "Success", detailed_info)
                playsound("voice/add_contact_voice.mp3")
            elif result['Status'] == 'Existing':
                QMessageBox.warning(self, "Already Exists", f"Contact {name} already exists.")
        except Exception as e:
            # 处理可能发生的任何异常
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")

        # 无论添加联系人成功与否，都清空输入框
        self.clearContactInputs()

    def clearContactInputs(self):
        # 清空所有输入框
        self.nameInput.clear()
        self.birthdayInput.clear()
        self.phoneInput.clear()
        self.emailInput.clear()
        self.heightInput.clear()
        self.weightInput.clear()
        self.collegeInput.clear()
        self.gradeInput.clear()
        self.majorInput.clear()
        self.gpaInput.clear()
        self.companyInput.clear()
        self.incomeInput.clear()
        self.acquaintanceInfoInput.clear()
        self.relationshipInput.clear()

    def initSearchContactsTab(self):
        self.searchContactsTab = QWidget()
        self.tabWidget.addTab(self.searchContactsTab, "Search Contacts")
        mainLayout = QVBoxLayout()  # 主布局
        layout = QVBoxLayout()

        # 搜索框
        self.searchInput = QLineEdit()
        layout.addWidget(QLabel("Enter Search Info:"))
        layout.addWidget(self.searchInput)

        # 搜索按钮
        searchBtn = QPushButton("Search")
        searchBtn.clicked.connect(self.searchContacts)
        layout.addWidget(searchBtn)

        # 创建表格显示搜索结果
        self.searchResultsTable = QTableWidget()
        layout.addWidget(self.searchResultsTable)  # 添加到水平布局
        layout.setAlignment(Qt.AlignCenter)  # 设置水平布局的对齐方式为居中

        # 调整最后一列宽度以填充剩余空间
        self.searchResultsTable.horizontalHeader().setStretchLastSection(True)
        mainLayout.addLayout(layout)  # 将水平布局添加到主布局

        self.searchContactsTab.setLayout(mainLayout)
    def searchContacts(self):
        search_info = self.searchInput.text()
        result = search_contact(search_info)

        # 清空当前的搜索结果表格
        self.searchResultsTable.setRowCount(0)
        self.searchResultsTable.setColumnCount(0)

        # 显示搜索结果
        if isinstance(result, dict) and result['Status'] == f'{len(result["Details"])} Contacts Found':
            # 设置表格列数和列标题
            headers = ["Name", "Birthday", "Phone", "Email", "Height", "Weight", "Contact Type", "Additional Information", "         "]
            self.searchResultsTable.setColumnCount(len(headers))
            self.searchResultsTable.setHorizontalHeaderLabels(headers)

            # 填充表格数据
            for contact in result['Details']:
                row_position = self.searchResultsTable.rowCount()
                self.searchResultsTable.insertRow(row_position)
                for col_num, item in enumerate(contact):
                    self.searchResultsTable.setItem(row_position, col_num, QTableWidgetItem(str(item)))

            # 调整列宽以适应内容
            self.searchResultsTable.resizeColumnsToContents()
            self.searchResultsTable.resizeRowsToContents()
        else:
            # 清空表格并显示消息
            self.searchResultsTable.setRowCount(1)
            self.searchResultsTable.setColumnCount(1)
            self.searchResultsTable.setItem(0, 0, QTableWidgetItem("No results to display"))

    def initModifyContactsTab(self):
        self.modifyContactsTab = QWidget()
        self.tabWidget.addTab(self.modifyContactsTab, "Modify Contacts")

        layout = QVBoxLayout()

        # 输入框：输入要修改的联系人姓名
        self.contactNameInput = QLineEdit()
        layout.addWidget(QLabel("Contact Name:"))
        layout.addWidget(self.contactNameInput)

        # Radio Buttons：选择要修改的信息类型
        self.modifyTypeGroup = QGroupBox("Information Type")
        modifyTypeLayout = QHBoxLayout()
        self.birthdayRadio = QRadioButton("Birthday")
        self.phoneRadio = QRadioButton("Phone")
        self.emailRadio = QRadioButton("Email")
        self.heightRadio = QRadioButton("Height")
        self.weightRadio = QRadioButton("Weight")
        modifyTypeLayout.addWidget(self.birthdayRadio)
        modifyTypeLayout.addWidget(self.phoneRadio)
        modifyTypeLayout.addWidget(self.emailRadio)
        modifyTypeLayout.addWidget(self.heightRadio)
        modifyTypeLayout.addWidget(self.weightRadio)
        self.modifyTypeGroup.setLayout(modifyTypeLayout)
        layout.addWidget(self.modifyTypeGroup)

        # 输入框：输入新的信息值
        self.newValueInput = QLineEdit()
        layout.addWidget(QLabel("New Value:"))
        layout.addWidget(self.newValueInput)

        # 修改按钮
        modifyBtn = QPushButton("Modify Contact")
        modifyBtn.clicked.connect(self.modifyContact)
        layout.addWidget(modifyBtn)

        self.modifyContactsTab.setLayout(layout)

    def modifyContact(self):
        name = self.contactNameInput.text()
        info_type, new_value = self.getSelectedInfoTypeAndValue()

        if not name or not info_type or not new_value:
            QMessageBox.warning(self, "Input Error", "Please complete all required fields.")
            return

        # 调用 modify_contact 函数
        result = modify_contact(name, info_type, new_value)
        playsound("voice/modify_voice.mp3")
        QMessageBox.information(self, "Result", result)

    def getSelectedInfoTypeAndValue(self):
        # 获取被选中的信息类型和对应的新值
        if self.birthdayRadio.isChecked():
            return 'birthday', self.newValueInput.text()
        elif self.phoneRadio.isChecked():
            return 'phone', self.newValueInput.text()
        elif self.emailRadio.isChecked():
            return 'email', self.newValueInput.text()
        elif self.heightRadio.isChecked():
            return 'height', self.newValueInput.text()
        elif self.weightRadio.isChecked():
            return 'weight', self.newValueInput.text()
        return None, None

    def initDeleteContactsTab(self):
        self.deleteContactsTab = QWidget()
        self.tabWidget.addTab(self.deleteContactsTab, "Delete Contacts")

        layout = QVBoxLayout()

        # 输入框：输入要删除的联系人姓名
        self.contactToDeleteInput = QLineEdit()
        layout.addWidget(QLabel("Contact Name to Delete:"))
        layout.addWidget(self.contactToDeleteInput)

        # 删除按钮
        deleteBtn = QPushButton("Delete Contact")
        deleteBtn.clicked.connect(self.deleteContact)
        layout.addWidget(deleteBtn)

        # Add new button to clear all contacts
        clearAllBtn = QPushButton("Clear All Contacts")
        clearAllBtn.clicked.connect(self.clearContacts)
        layout.addWidget(clearAllBtn)

        self.deleteContactsTab.setLayout(layout)


    def deleteContact(self):
        name = self.contactToDeleteInput.text()

        if not name:
            QMessageBox.warning(self, "Input Error", "Please enter a contact name.")
            return

        # 调用 delete_contact 函数
        result = delete_contact(name)
        QMessageBox.information(self, "Result", result)
        if result == "Contact deleted successfully.":
            playsound("voice/delete_voice.mp3")
            self.contactToDeleteInput.clear()

    def clearContacts(self):
        # Confirmation dialog
        reply = QMessageBox.question(self, 'Confirm Clear',
                                     "Are you sure you want to clear all contacts? This action cannot be undone.",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            # User confirmed the action
            try:
                clear()  # Assuming clear() is properly imported
                QMessageBox.information(self, "Success", "All contacts cleared successfully.")
                playsound("voice/delete_voice.mp3")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")
        else:
            # User declined the action
            QMessageBox.information(self, "Cancelled", "Operation cancelled.")

    def initSortContactsTab(self):
        self.sortContactsTab = QWidget()
        self.tabWidget.addTab(self.sortContactsTab, "Sort Contacts")

        layout = QVBoxLayout()

        # 排序类型选择
        self.sortTypeGroup = QGroupBox("Sort Type")
        sortTypeLayout = QHBoxLayout()
        self.generalSortRadio = QRadioButton("General Sort")
        self.studentSortRadio = QRadioButton("Student Sort")
        self.colleagueSortRadio = QRadioButton("Colleague Sort")
        sortTypeLayout.addWidget(self.generalSortRadio)
        sortTypeLayout.addWidget(self.studentSortRadio)
        sortTypeLayout.addWidget(self.colleagueSortRadio)
        self.sortTypeGroup.setLayout(sortTypeLayout)
        layout.addWidget(self.sortTypeGroup)

        # 普通排序的字段选择
        self.generalSortOptionsGroup = QGroupBox("Sort Options - General")
        generalSortOptionsLayout = QHBoxLayout()
        self.birthdaySortRadio = QRadioButton("Birthday")
        self.heightSortRadio = QRadioButton("Height")
        self.weightSortRadio = QRadioButton("Weight")
        self.phoneSortRadio = QRadioButton("Phone")
        generalSortOptionsLayout.addWidget(self.birthdaySortRadio)
        generalSortOptionsLayout.addWidget(self.heightSortRadio)
        generalSortOptionsLayout.addWidget(self.weightSortRadio)
        generalSortOptionsLayout.addWidget(self.phoneSortRadio)
        self.generalSortOptionsGroup.setLayout(generalSortOptionsLayout)
        layout.addWidget(self.generalSortOptionsGroup)

        # 学生排序的字段选择
        self.studentSortOptionsGroup = QGroupBox("Sort Options - Student")
        studentSortOptionsLayout = QHBoxLayout()
        self.gradeSortRadio = QRadioButton("Grade")
        self.gpaSortRadio = QRadioButton("GPA")
        studentSortOptionsLayout.addWidget(self.gradeSortRadio)
        studentSortOptionsLayout.addWidget(self.gpaSortRadio)
        self.studentSortOptionsGroup.setLayout(studentSortOptionsLayout)
        layout.addWidget(self.studentSortOptionsGroup)

        # 排序按钮
        sortBtn = QPushButton("Sort")
        sortBtn.clicked.connect(self.sortContacts)
        layout.addWidget(sortBtn)

        # 结果展示区域
        self.sortResultsTable = QTableWidget()
        layout.addWidget(self.sortResultsTable)

        self.sortContactsTab.setLayout(layout)
    def sortContacts(self):
        # 检查用户选择的排序类型
        if self.generalSortRadio.isChecked():
            sort_key = self.getSelectedGeneralSortKey()
            if sort_key:
                sorted_contacts = sort(sort_key)
                self.displaySortResults(sorted_contacts, sort_key)

        elif self.studentSortRadio.isChecked():
            sort_key = 'grade' if self.gradeSortRadio.isChecked() else 'gpa'
            sorted_students = sort_students(sort_key)
            self.displaySortResults(sorted_students, sort_key)

        elif self.colleagueSortRadio.isChecked():
            sorted_colleagues = sort_colleagues_by_income()
            self.displaySortResults(sorted_colleagues, "income")

    def getSelectedGeneralSortKey(self):
        # 获取普通排序选择的键
        if self.birthdaySortRadio.isChecked():
            return 'birthday'
        elif self.heightSortRadio.isChecked():
            return 'height'
        elif self.weightSortRadio.isChecked():
            return 'weight'
        elif self.phoneSortRadio.isChecked():
            return 'phone'
        return None

    def displaySortResults1(self, results):
        # 显示排序结果
        self.sortResults.clear()
        if isinstance(results, list):
            playsound("voice/sortall.mp3")
            for contact in results:
                self.sortResults.append(', '.join(str(item) for item in contact))
        else:
            self.sortResults.setText(results)

    def displaySortResults2(self, results):
        # 显示排序结果
        self.sortResults.clear()
        if isinstance(results, list):
            playsound("voice/sortall.mp3")
            for contact in results:
                # 确保每个键都存在于字典中
                name = contact.get('name', 'Unknown')
                birthday = contact.get('birthday', 'N/A')
                phone = contact.get('phone', 'N/A')
                email = contact.get('email', 'N/A')
                height = contact.get('height', 'N/A')
                weight = contact.get('weight', 'N/A')

                # 格式化输出
                # playsound("voice/sort.mp3")
                contact_info = f"Name: {name}, Birthday: {birthday}, Phone: {phone}, Email: {email}, Height: {height}, Weight: {weight}"
                self.sortResults.append(contact_info)
        else:
            self.sortResults.setText(results)

    def displaySortResults(self, results, sort_key):
        # 清空现有的表格数据
        self.sortResultsTable.setRowCount(0)
        self.sortResultsTable.setColumnCount(0)

        if isinstance(results, list) and results:
            playsound("voice/sortall.mp3")

            # 假设结果是字典列表，获取列标题
            headers = list(results[0].keys()) + ["   "] if isinstance(results[0], dict) else ["Name",
                                                                                              sort_key.capitalize(),
                                                                                              "   "]
            self.sortResultsTable.setColumnCount(len(headers))
            self.sortResultsTable.setHorizontalHeaderLabels(headers)

            # 填充表格数据
            for contact in results:
                row_position = self.sortResultsTable.rowCount()
                self.sortResultsTable.insertRow(row_position)
                if isinstance(contact, dict):
                    # 从字典中获取数据
                    for col_num, key in enumerate(headers):
                        self.sortResultsTable.setItem(row_position, col_num,
                                                      QTableWidgetItem(str(contact.get(key, 'N/A'))))
                else:
                    # 如果数据是列表或元组，请确保按照列标题的顺序来访问值
                    for col_num, item in enumerate(contact):
                        self.sortResultsTable.setItem(row_position, col_num, QTableWidgetItem(str(item)))


            # 调整列宽以适应内容，并使最后一列自动扩展
            self.sortResultsTable.horizontalHeader().setStretchLastSection(True)

            # 调整列宽以适应内容
            self.sortResultsTable.resizeColumnsToContents()
            self.sortResultsTable.resizeRowsToContents()

        else:
            # 如果没有结果或结果不是预期的列表，显示消息
            self.sortResultsTable.setRowCount(1)
            self.sortResultsTable.setColumnCount(1)
            self.sortResultsTable.setItem(0, 0, QTableWidgetItem("No results to display"))

    def initFilterContactsTab(self):
        self.filterContactsTab = QWidget()
        self.tabWidget.addTab(self.filterContactsTab, "Filter Contacts")

        layout = QVBoxLayout()

        # 输入框：输入要筛选的联系人姓名
        self.filterNameInput = QLineEdit()
        layout.addWidget(QLabel("Filter by Name:"))
        layout.addWidget(self.filterNameInput)

        # Radio Buttons：选择要筛选的联系人类型
        self.filterTypeGroup = QGroupBox("Filter by Contact Type")
        filterTypeLayout = QHBoxLayout()
        self.studentFilterRadio = QRadioButton("Student")
        self.colleagueFilterRadio = QRadioButton("Colleague")
        self.friendFilterRadio = QRadioButton("Friend")
        self.relativeFilterRadio = QRadioButton("Relative")
        filterTypeLayout.addWidget(self.studentFilterRadio)
        filterTypeLayout.addWidget(self.colleagueFilterRadio)
        filterTypeLayout.addWidget(self.friendFilterRadio)
        filterTypeLayout.addWidget(self.relativeFilterRadio)
        self.filterTypeGroup.setLayout(filterTypeLayout)
        layout.addWidget(self.filterTypeGroup)

        # 筛选按钮
        filterBtn = QPushButton("Filter Contacts")
        filterBtn.clicked.connect(self.filterContacts)
        layout.addWidget(filterBtn)

        # 创建表格显示筛选结果
        self.filterResultsTable = QTableWidget()
        layout.addWidget(self.filterResultsTable)

        self.filterContactsTab.setLayout(layout)

        # 高级筛选选项 - 身高和体重
        self.advancedFilterGroup = QGroupBox("Advanced Filter Options")
        advancedFilterLayout = QFormLayout()
        self.heightFilterInput = QLineEdit()
        self.weightFilterInput = QLineEdit()
        self.heightFilterOperator = QComboBox()
        self.weightFilterOperator = QComboBox()
        self.heightFilterOperator.addItems(['>', '<', '='])
        self.weightFilterOperator.addItems(['>', '<', '='])
        advancedFilterLayout.addRow(QLabel("Height Filter:"), self.heightFilterOperator)
        advancedFilterLayout.addRow(self.heightFilterInput)
        advancedFilterLayout.addRow(QLabel("Weight Filter:"), self.weightFilterOperator)
        advancedFilterLayout.addRow(self.weightFilterInput)
        self.advancedFilterGroup.setLayout(advancedFilterLayout)
        layout.addWidget(self.advancedFilterGroup)

    def filterContacts(self):
        """
        根据用户输入的筛选条件来过滤联系人列表。
        该方法首先从文本输入框获取名字和联系人类型的筛选条件。
        然后，根据用户选择的高级筛选条件（身高和体重）构建更详细的筛选标准。
        最后，使用这些标准调用高级筛选功能，并展示筛选结果。

        高级筛选功能允许用户根据身高和体重的特定范围来过滤联系人。
        """

        # 从输入框获取用户输入的名字
        name = self.filterNameInput.text()

        # 获取用户选中的联系人类型
        contact_type = self.getSelectedFilterType()

        # 初始化筛选标准字典
        criteria = {}

        # 如果用户输入了名字，添加到筛选标准中
        if name:
            criteria['Name'] = name

        # 如果用户选择了联系人类型，添加到筛选标准中
        if contact_type:
            criteria['Contact Type'] = contact_type

        # 获取用户输入的高级筛选标准：身高
        height_criteria = self.getAdvancedFilterCriteria(self.heightFilterOperator.currentText(),
                                                         self.heightFilterInput.text())

        # 获取用户输入的高级筛选标准：体重
        weight_criteria = self.getAdvancedFilterCriteria(self.weightFilterOperator.currentText(),
                                                         self.weightFilterInput.text())

        # 如果用户输入了高级筛选标准，添加到筛选标准中
        if height_criteria is not None:
            criteria['Height'] = height_criteria
        if weight_criteria is not None:
            criteria['Weight'] = weight_criteria

        # 调用高级筛选函数，传入所有筛选标准
        filtered_contacts = filter_contacts_advanced(criteria)

        # 显示筛选后的联系人结果
        # playsound("voice/filter_voice.mp3")
        self.displayFilterResults(filtered_contacts)

    def getAdvancedFilterCriteria(self, operator, value):
        # 解析高级筛选条件
        if value:
            try:
                value = float(value)
                return (operator, value)
            except ValueError:
                return None
        return None

    def getSelectedFilterType(self):
        # 获取选中的联系人类型
        if self.studentFilterRadio.isChecked():
            return 'student'
        elif self.colleagueFilterRadio.isChecked():
            return 'colleague'
        elif self.friendFilterRadio.isChecked():
            return 'friend'
        elif self.relativeFilterRadio.isChecked():
            return 'relative'
        return None

    def displayFilterResults(self, results):
        # 清空现有的表格数据
        self.filterResultsTable.setRowCount(0)
        self.filterResultsTable.setColumnCount(0)

        if isinstance(results, list) and results:
            playsound("voice/filter_voice.mp3")

            # 确定表格的列数
            column_count = len(results[0])
            self.filterResultsTable.setColumnCount(column_count)

            # 为表格设置列标题
            headers = ["Name", "Birthday", "Phone", "Email", "Height", "Weight", "Contact Type","Additional Information ","   "]
            self.filterResultsTable.setHorizontalHeaderLabels(headers)

            # 填充表格数据
            for contact in results:
                row_position = self.filterResultsTable.rowCount()
                self.filterResultsTable.insertRow(row_position)
                for col_num, item in enumerate(contact):
                    self.filterResultsTable.setItem(row_position, col_num, QTableWidgetItem(str(item)))

            # 调整列宽以适应内容，并使最后一列自动扩展
            self.filterResultsTable.horizontalHeader().setStretchLastSection(True)

            # 调整其他列的宽度以适应内容
            for i in range(column_count - 1):  # 调整除最后一列之外的所有列
                self.filterResultsTable.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)

        else:
            # 如果没有结果或结果不是列表，显示消息
            self.filterResultsTable.setRowCount(1)
            self.filterResultsTable.setColumnCount(1)
            self.filterResultsTable.setItem(0, 0, QTableWidgetItem("No contacts found."))

    # Define methods to connect UI actions to your backend logic
    def initFigueContactsTab(self):
        self.figueContactsTab = QWidget()
        self.tabWidget.addTab(self.figueContactsTab, "Figue Contacts")

        layout = QVBoxLayout()

        # 创建 Matplotlib 图表
        fig, ax = plt.subplots()
        self.canvas = FigureCanvas(fig)
        self.plotContactsPieChart(ax)

        layout.addWidget(self.canvas)
        self.figueContactsTab.setLayout(layout)

    def plotContactsPieChart(self, ax):
        try:
            contact_types, numbers = self.getContactTypesData()

            # 检查是否有数据来绘制饼图
            if not contact_types or not numbers:
                ax.clear()  # 清除之前的绘图
                ax.text(0.5, 0.5, '没有可用的联系人数据', ha='center', va='center')
                ax.set_title("通讯录关系占比")
                self.canvas.draw()
                return

            plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
            ax.pie(numbers, labels=contact_types, autopct="(%1.1f%%)", startangle=90)
            ax.set_title("通讯录关系占比")
            self.canvas.draw()
        except FileNotFoundError:
            print("contacts.xlsx file not found.")

    def getContactTypesData(self):
        """
        从 Excel 文件中提取联系人类型的数据，并进行统计。
        该方法首先加载一个名为 'contacts.xlsx' 的 Excel 文件。
        然后遍历 'Contacts' 工作表中的每一行，以统计不同联系人类型的数量。
        最后，返回一个排序后的联系人类型及其计数的列表。

        Returns:
            zip: 一个包含两个元素的元组列表，分别是联系人类型和对应的数量。
                 列表是按照联系人数量降序排序的。
        """

        # 指定要读取的 Excel 文件名
        filename = "contacts.xlsx"
        try:
            # 使用 openpyxl 库加载工作簿
            wb = openpyxl.load_workbook(filename)

            # 选择工作表 'Contacts'
            sheet = wb['Contacts']

            # 初始化一个字典来存储不同联系人类型的计数
            dict_contact = {}

            # 遍历工作表中的每一行
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                # 提取每行的联系人类型（假设存储在第七列，即索引 6 的位置）
                contact_type = row[6]

                # 如果这个类型的联系人尚未在字典中，将其添加到字典中，并设置计数为 1
                if contact_type not in dict_contact:
                    dict_contact[contact_type] = 1
                # 如果这个类型的联系人已经在字典中，将其计数增加 1
                else:
                    dict_contact[contact_type] += 1

            # 将字典中的项按照数量进行降序排序，并返回一个 zip 对象
            # zip 对象包含了排序后的类型和数量，可以直接用于迭代或转换为列表
            return zip(*sorted(dict_contact.items(), key=lambda item: item[1], reverse=True))

        except FileNotFoundError:
            print("contacts.xlsx file not found.")
            return [], []


    def initSendBirthdayEmailsTab(self):
        self.sendBirthdayEmailsTab = QWidget()
        self.tabWidget.addTab(self.sendBirthdayEmailsTab, "Send Birthday Emails")

        layout = QVBoxLayout()
        splitter = QSplitter(Qt.Vertical)  # 创建一个垂直分割器

        # 添加一个表格来显示即将过生日的人的信息
        self.birthdayTable = QTableWidget()
        self.birthdayTable.setColumnCount(2)  # 名字和邮箱
        self.birthdayTable.setHorizontalHeaderLabels(["Name", "Email"])
        self.birthdayTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)  # 列宽自适应
        layout.addWidget(self.birthdayTable)

        # 创建并设置文本区域来显示邮件发送结果
        self.emailResults = QTextEdit()
        self.emailResults.setReadOnly(True)
        font = QFont("Arial", 10)  # 设置字体和大小
        self.emailResults.setFont(font)
        splitter.addWidget(self.emailResults)

        layout.addWidget(splitter)
        self.sendBirthdayEmailsTab.setLayout(layout)

        # 当 UI 启动时自动发送生日邮件
        self.autoSendBirthdayEmails()

    birthday_wishes = [
        "祝您生日快乐，岁岁平安！",
        "在这特别的日子里，愿您的笑容如阳光般灿烂！",
        "生日快乐！愿您的每一天都充满幸福和喜悦。",
        "愿这美妙的一天带给您无尽的快乐和美好的回忆！",
        "又长大一岁了，祝您越来越年轻，生日快乐！",
        "愿您的生日充满无尽的快乐，愿您的每天都闪耀着快乐的阳光。",
        "在您的生日之际，向您致以最美好的祝愿！",
        "愿您在这特别的日子里，心想事成，万事如意！"
    ]
    def autoSendBirthdayEmails(self):
        filename = "contacts.xlsx"
        birthdays_tomorrow = get_tomorrows_birthdays(filename)

        if birthdays_tomorrow:
            playsound("voice/yes_birthday.mp3")
            self.birthdayTable.setRowCount(len(birthdays_tomorrow))
            for i, (email, name) in enumerate(birthdays_tomorrow):
                self.birthdayTable.setItem(i, 0, QTableWidgetItem(name))
                self.birthdayTable.setItem(i, 1, QTableWidgetItem(email))

                # 生成随机的祝福语
                birthday_message = random.choice(birthday_wishes)

                try:
                    # 发送邮件，使用生成的祝福语
                    send_birthday_email(email, name, birthday_message)
                    # 在 UI 中显示相同的祝福语
                    self.emailResults.append(f"生日祝福邮件已发送至 {email}：\n{name}，{birthday_message}\n")
                except Exception as e:
                    self.emailResults.append(f"邮件发送失败至 {email}: {str(e)}")
        else:
            playsound("voice/no_birthday.mp3")
            self.emailResults.setText("明天没有人过生日。")

    def initDisplayAllContactsTab(self):
        self.displayAllContactsTab = QWidget()
        self.tabWidget.addTab(self.displayAllContactsTab, "Display All Contacts")

        layout = QVBoxLayout()
        splitter = QSplitter()  # 使用分割器来分隔表格和图表

        # 创建表格显示所有联系人
        self.contactsTable = QTableWidget()
        self.populateContactsTable()
        splitter.addWidget(self.contactsTable)

        # 创建 Matplotlib 图表
        fig, ax = plt.subplots()
        self.canvas = FigureCanvas(fig)
        self.plotContactsPieChart(ax)
        splitter.addWidget(self.canvas)

        # 设置分割器中控件的初始大小
        splitter.setSizes([400, 200])  # 例如，表格部分 400，图表部分 200

        layout.addWidget(splitter)
        self.displayAllContactsTab.setLayout(layout)

    def populateContactsTable(self):
        """
        从后端获取所有联系人数据，并将其填充到表格中。
        该方法首先调用 get_all_contacts 函数来获取所有联系人的数据。
        然后，根据获取到的数据设置表格的行和列。
        最后，将每个联系人的数据填充到表格中的相应位置。
        如果没有获取到数据，表格将被清空。
        """

        # 从后端获取所有联系人的数据
        all_contacts = get_all_contacts()

        # 检查获取到的数据是否为列表且不为空
        if isinstance(all_contacts, list) and all_contacts:
            # 设置表格的行数和列数
            self.contactsTable.setRowCount(len(all_contacts))
            self.contactsTable.setColumnCount(len(all_contacts[0]))

            # 设置表格的水平标题
            headers = ["Name", "Birthday", "Phone", "Email", "Height", "Weight", "Contact Type"]
            self.contactsTable.setHorizontalHeaderLabels(headers)

            # 遍历所有联系人数据，将其添加到表格中
            for row_num, row_data in enumerate(all_contacts):
                for col_num, col_data in enumerate(row_data):
                    # 将每个数据项添加到表格的相应位置
                    self.contactsTable.setItem(row_num, col_num, QTableWidgetItem(str(col_data)))
        else:
            # 如果没有获取到联系人数据，清空表格
            self.contactsTable.setRowCount(0)
            self.contactsTable.setColumnCount(0)
            # 打印获取到的数据，用于调试
            print(all_contacts)


def main():
    app = QApplication(sys.argv)
    ex = ContactManagementApp()
    ex.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
